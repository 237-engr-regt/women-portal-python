from flask import Flask, render_template, request, jsonify, send_file, redirect
import os
from dotenv import load_dotenv
import random
import base64
from openpyxl import Workbook, load_workbook
from threading import Lock
from datetime import datetime

# ================= INIT =================
load_dotenv()

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = os.getenv("SECRET_KEY", "secret123")

print("🔥 FINAL COMPLETE SYSTEM RUNNING")

# ================= PATH =================
UPLOAD_FOLDER = "audio_files"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ❗ IMPORTANT: /tmp hata diya (permanent save)
EXCEL_FILE = "complaints.xlsx"

# ================= EXCEL =================
excel_lock = Lock()

def create_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "ID", "Complaint ID", "Name", "Address", "Contact",
            "Email", "Unit", "WO", "Quarter", "Complaint",
            "Category", "Subcategory", "Reply", "Audio", "Date"
        ])
        wb.save(EXCEL_FILE)

create_excel()

def save_to_excel(data):
    with excel_lock:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws.append([
            ws.max_row,
            data.get("complaint_id"),
            data.get("name"),
            data.get("address"),
            data.get("contact"),
            data.get("email"),
            data.get("unit"),
            data.get("wo"),
            data.get("quarter"),
            data.get("complaint"),
            data.get("category"),
            data.get("subcategory"),
            data.get("reply"),
            data.get("audio"),
            now
        ])
        wb.save(EXCEL_FILE)

# ================= ROUTES =================

@app.route('/')
def home():
    return render_template("landing.html")

# ✅ COMPLAINT
@app.route('/complaint', methods=['GET', 'POST'])
def complaint():

    if request.method == 'POST':
        try:
            complaint_id = "CMP" + str(random.randint(10000, 99999))

            data = {
                "complaint_id": complaint_id,
                "name": request.form.get('name', ''),
                "address": request.form.get('address', ''),
                "contact": request.form.get('contact', ''),
                "email": request.form.get('email', ''),
                "unit": request.form.get('unit', ''),
                "wo": request.form.get('wo', ''),
                "quarter": request.form.get('quarter', ''),
                "complaint": request.form.get('complaint', ''),
                "category": request.form.get('category', ''),
                "subcategory": request.form.get('subcategory', ''),
                "reply": "Pending",
                "audio": ""
            }

            # 🎤 AUDIO SAVE
            audio_data = request.form.get("audio_data")
            if audio_data and "," in audio_data:
                try:
                    header, encoded = audio_data.split(",", 1)
                    file_data = base64.b64decode(encoded)
                    filepath = os.path.join(UPLOAD_FOLDER, f"{complaint_id}.webm")

                    with open(filepath, "wb") as f:
                        f.write(file_data)

                    data["audio"] = filepath
                except Exception as e:
                    print("⚠️ Audio error:", e)

            save_to_excel(data)

            return jsonify({"status": "success", "id": complaint_id})

        except Exception as e:
            print("❌ ERROR:", e)
            return jsonify({"status": "error", "message": str(e)})

    return render_template("complaint.html")


# ✅ TRACK COMPLAINT (NEW 🔥)
@app.route('/track', methods=['GET', 'POST'])
def track():
    data = None

    if request.method == 'POST':
        cid = request.form.get("complaint_id")

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[1]) == cid:
                data = row
                break

    return render_template("track.html", data=data)


# ✅ ADMIN PANEL (NEW 🔥)
@app.route('/admin')
def admin():
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        data = list(ws.values)

        return render_template("admin.html", data=data)

    except Exception as e:
        print("❌ Admin Error:", e)
        return "Admin Error"


# ✅ DOWNLOAD EXCEL
@app.route("/download_excel")
def download_excel():
    try:
        return send_file(EXCEL_FILE, as_attachment=True)
    except Exception as e:
        print("❌ Download Error:", e)
        return "File not found"


# ✅ REPLY SYSTEM
@app.route('/reply/<cid>', methods=['POST'])
def reply(cid):
    try:
        reply_text = request.form.get("reply", "")

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if str(row[1].value) == cid:
                row[12].value = reply_text
                break

        wb.save(EXCEL_FILE)

        return redirect("/admin")

    except Exception as e:
        print("❌ Reply Error:", e)
        return "Reply Error"


# ================= RUN =================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)